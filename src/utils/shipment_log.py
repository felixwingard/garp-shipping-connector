"""Sändningslogg — sparar alla sändningar till Excel för enkel spårning.

En Excel-fil (sandningslogg.xlsx) i programmets rotmapp som uppdateras
vid varje lyckad bokning. Kan öppnas direkt i Excel för sökning.
"""

import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional

from .config import get_base_dir

logger = logging.getLogger(__name__)

_LOCK = threading.Lock()

_HEADERS = [
    "Datum",
    "Ordernr",
    "Spårningsnummer",
    "Spårningslänk",
    "Transportör",
    "Produkt",
    "Mottagare",
    "Adress",
    "Postnummer",
    "Stad",
    "Land",
    "Vikt (kg)",
    "Kolli",
    "E-post",
    "Telefon",
    "Pris",
]

_COL_WIDTHS = {
    "A": 18, "B": 18, "C": 26, "D": 34, "E": 12, "F": 10,
    "G": 24, "H": 28, "I": 12, "J": 16, "K": 8,
    "L": 10, "M": 8, "N": 24, "O": 16, "P": 12,
}

_TRACKING_URLS = {
    "DHL": "https://www.dhl.com/se-sv/home/tracking.html?tracking-id={tracking}",
    "Bring": "https://tracking.bring.se/tracking/{tracking}",
}


def _log_path() -> Path:
    return get_base_dir() / "sandningslogg.xlsx"


def append_shipment(
    order_no: str,
    tracking: str,
    carrier: str,
    product_code: str,
    receiver_name: str = "",
    receiver_address: str = "",
    receiver_zipcode: str = "",
    receiver_city: str = "",
    receiver_country: str = "",
    receiver_email: str = "",
    receiver_phone: str = "",
    weight_kg: float = 0,
    copies: int = 1,
    estimated_price: str = "",
) -> None:
    """Lägger till en sändning i sändningsloggen (Excel)."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl saknas — kör 'pip install openpyxl' för sändningslogg")
        return

    with _LOCK:
        path = _log_path()
        datestr = datetime.now().strftime("%Y-%m-%d %H:%M")

        tracking_url = ""
        carrier_upper = (carrier or "").upper()
        for key, template in _TRACKING_URLS.items():
            if key.upper() in carrier_upper:
                tracking_url = template.format(tracking=tracking)
                break

        row_data = [
            datestr,
            order_no,
            tracking,
            "Spåra försändelse",
            carrier,
            product_code,
            receiver_name,
            receiver_address,
            receiver_zipcode,
            receiver_city,
            receiver_country,
            weight_kg if weight_kg else "",
            copies if copies else "",
            receiver_email,
            receiver_phone,
            estimated_price,
        ]

        try:
            if path.exists():
                wb = load_workbook(path, read_only=False)
                ws = wb.active
                ws.append(row_data)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sändningar"
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_text = Font(bold=True, size=11, color="FFFFFF")
                for col, h in enumerate(_HEADERS, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = header_text
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for letter, width in _COL_WIDTHS.items():
                    ws.column_dimensions[letter].width = width
                ws.auto_filter.ref = f"A1:P1"
                ws.freeze_panes = "A2"
                ws.append(row_data)

            if tracking_url:
                link_cell = ws.cell(row=ws.max_row, column=4)
                link_cell.hyperlink = tracking_url
                link_cell.font = Font(color="0563C1", underline="single")

            wb.save(path)
            wb.close()
        except Exception as e:
            logger.warning(f"Kunde inte skriva sändningslogg: {e}")
