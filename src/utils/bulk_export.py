"""Bring bulk — spårning av ordrar i Excel och backup-export."""

import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from .config import get_base_dir

logger = logging.getLogger(__name__)

_LOCK = threading.Lock()

_ORDERS_DIR = "bulk_orders"
_EXPORTS_DIR = "bulk_exports"

_ORDER_HEADERS = [
    "Datum", "Ordernr", "Spårningsnummer", "Mottagare", "Adress", "Postnummer", "Stad", "Land",
    "Vikt (kg)", "Kolli",
]


def _orders_dir() -> Path:
    d = get_base_dir() / _ORDERS_DIR
    d.mkdir(parents=True, exist_ok=True)
    return d


def _exports_dir() -> Path:
    d = get_base_dir() / _EXPORTS_DIR
    d.mkdir(parents=True, exist_ok=True)
    return d


def _excel_path_for_bulk(bulk_id: str) -> Path:
    safe = bulk_id.replace("/", "-").replace("\\", "-").strip()
    return _orders_dir() / f"bulk_{safe}.xlsx"


def append_bring_bulk_order(
    bulk_id: str,
    order_no: str,
    tracking: str,
    weight_kg: float,
    kolli: int,
    receiver_name: str = "",
    receiver_address: str = "",
    receiver_zipcode: str = "",
    receiver_city: str = "",
    receiver_country: str = "",
) -> None:
    """Lägger till en order i bulk-Excel-filen (anropas vid varje lyckad Bring bulk-bokning)."""
    if not bulk_id or not bulk_id.strip():
        return
    bulk_id = bulk_id.strip()
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        logger.warning("openpyxl saknas — bulk-ordrar sparas ej")
        return

    with _LOCK:
        path = _excel_path_for_bulk(bulk_id)
        datestr = datetime.now().strftime("%Y-%m-%d %H:%M")
        row_data = [
            datestr,
            order_no,
            tracking,
            receiver_name,
            receiver_address,
            receiver_zipcode,
            receiver_city,
            receiver_country,
            weight_kg,
            kolli,
        ]
        if path.exists():
            try:
                wb = load_workbook(path, read_only=False)
                ws = wb.active
                ws.append(row_data)
                wb.save(path)
                wb.close()
            except Exception as e:
                logger.warning(f"Kunde inte lägga till i bulk Excel: {e}")
        else:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Ordrar"
                for col, h in enumerate(_ORDER_HEADERS, 1):
                    ws.cell(row=1, column=col, value=h)
                ws.append(row_data)
                wb.save(path)
                wb.close()
            except Exception as e:
                logger.warning(f"Kunde inte skapa bulk Excel: {e}")


def get_bulk_orders(bulk_id: str) -> list[dict[str, Any]]:
    """Returnerar ordrar sparade för bulk_id (från Excel)."""
    bulk_id = (bulk_id or "").strip()
    if not bulk_id:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    with _LOCK:
        path = _excel_path_for_bulk(bulk_id)
        if not path.exists():
            return []
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            result = []
            for r in rows:
                if not r or (r[0] is None and (len(r) < 2 or r[1] is None)):
                    continue
                # Kolumnordning: Datum, Ordernr, Spårning, Mottagare, Adress, Postnr, Stad, Land, Vikt, Kolli
                # Bakåtkompatibel: gamla filer har 9 kolumner (utan Datum)
                if len(r) >= 10:
                    result.append({
                        "datum": r[0] or "",
                        "order_no": r[1] or "",
                        "tracking": r[2] or "",
                        "receiver_name": r[3] or "",
                        "receiver_address": r[4] or "",
                        "receiver_zipcode": r[5] or "",
                        "receiver_city": r[6] or "",
                        "receiver_country": r[7] or "",
                        "weight_kg": r[8] if r[8] is not None else 0,
                        "kolli": r[9] if len(r) > 9 and r[9] is not None else 1,
                    })
                else:
                    result.append({
                        "datum": "",
                        "order_no": r[0] or "",
                        "tracking": r[1] or "" if len(r) > 1 else "",
                        "receiver_name": r[2] or "" if len(r) > 2 else "",
                        "receiver_address": r[3] or "" if len(r) > 3 else "",
                        "receiver_zipcode": r[4] or "" if len(r) > 4 else "",
                        "receiver_city": r[5] or "" if len(r) > 5 else "",
                        "receiver_country": r[6] or "" if len(r) > 6 else "",
                        "weight_kg": r[7] if len(r) > 7 else 0,
                        "kolli": r[8] if len(r) > 8 and r[8] is not None else 1,
                    })
            return result
        except Exception as e:
            logger.warning(f"Kunde inte läsa bulk Excel: {e}")
            return []


def get_bulk_totals(bulk_id: str) -> dict[str, Any]:
    """Beräknar totalvikt och antal kolli från bulk-Excel. Returnerar {'total_weight_kg', 'num_packages'}."""
    orders = get_bulk_orders(bulk_id)
    total_weight = sum(float(o.get("weight_kg") or 0) for o in orders)
    num_packages = sum(int(o.get("kolli") or 1) for o in orders)
    return {"total_weight_kg": total_weight, "num_packages": num_packages}


def clear_bulk_orders(bulk_id: str) -> None:
    """Tar bort bulk-Excel-filen (efter export)."""
    bulk_id = (bulk_id or "").strip()
    if not bulk_id:
        return
    with _LOCK:
        path = _excel_path_for_bulk(bulk_id)
        try:
            path.unlink(missing_ok=True)
        except Exception as e:
            logger.warning(f"Kunde inte ta bort bulk Excel: {e}")


def export_bulk_to_excel(
    bulk_id: str,
    total_weight_kg: int,
    num_packages: int,
    num_pallets: int,
    num_direct_pallets: int = 0,
    direct_pallets_weight_kg: int = 0,
    num_invoices: int = 3,
    waybill_url: str = "",
    routing_url: str = "",
    orders: Optional[list[dict[str, Any]]] = None,
) -> Optional[Path]:
    """Skapar Excel-backup av bulk-sändningen. Returnerar sökväg till filen eller None vid fel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        logger.warning("openpyxl ej installerad — Excel-export hoppas över")
        return None

    orders = orders or []
    safe_bulk = bulk_id.replace("/", "-").replace("\\", "-").strip()
    datestr = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"Bring_Bulk_{safe_bulk}_{datestr}.xlsx"
    filepath = _exports_dir() / filename

    wb = Workbook()

    # --- Blad 1: Sammanfattning ---
    ws = wb.active
    ws.title = "Sammanfattning"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28

    row = 1
    ws.cell(row=row, column=1, value="Bring Bulk — Backup").font = Font(bold=True, size=14)
    row += 2
    rows_data = [
        ("Bulk-ID", bulk_id),
        ("Datum", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Totalvikt (kg)", total_weight_kg),
        ("Antal kolli (paket)", num_packages),
        ("Antal pall (Bulk)", num_pallets),
        ("Hel pall till kund (antal)", num_direct_pallets),
        ("Hel pall till kund (vikt kg)", direct_pallets_weight_kg),
        ("Antal fakturakopior", num_invoices),
    ]
    for label, val in rows_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1
    if waybill_url:
        ws.cell(row=row, column=1, value="CMR-waybill").font = Font(bold=True)
        ws.cell(row=row, column=2, value=waybill_url)
        row += 1
    if routing_url:
        ws.cell(row=row, column=1, value="Routing labels").font = Font(bold=True)
        ws.cell(row=row, column=2, value=routing_url)

    # --- Blad 2: Ordrar (med mottagare, adress etc.) ---
    ws2 = wb.create_sheet("Ordrar", 1)
    for col, h in enumerate(_ORDER_HEADERS, 1):
        ws2.cell(row=1, column=col, value=h).font = Font(bold=True)
    for i, o in enumerate(orders, 2):
        ws2.cell(row=i, column=1, value=o.get("datum", ""))
        ws2.cell(row=i, column=2, value=o.get("order_no", ""))
        ws2.cell(row=i, column=3, value=o.get("tracking", ""))
        ws2.cell(row=i, column=4, value=o.get("receiver_name", ""))
        ws2.cell(row=i, column=5, value=o.get("receiver_address", ""))
        ws2.cell(row=i, column=6, value=o.get("receiver_zipcode", ""))
        ws2.cell(row=i, column=7, value=o.get("receiver_city", ""))
        ws2.cell(row=i, column=8, value=o.get("receiver_country", ""))
        ws2.cell(row=i, column=9, value=o.get("weight_kg"))
        ws2.cell(row=i, column=10, value=o.get("kolli", 1))
    from openpyxl.utils import get_column_letter
    for col in range(1, 11):
        ws2.column_dimensions[get_column_letter(col)].width = 18 if col <= 2 else 14

    wb.save(filepath)
    logger.info(f"Bulk Excel-backup sparad: {filepath}")
    return filepath
